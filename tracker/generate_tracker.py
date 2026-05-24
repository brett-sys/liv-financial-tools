#!/usr/bin/env python3
"""Generate tracker/bet_tracker.xlsx — Brett's betting tracker.

Single source of truth for sizing lives on the Dashboard sheet:
    Bankroll  = Dashboard!$B$2
    Lean      = Dashboard!$B$4   ($295)
    Standard  = Dashboard!$B$5   ($510)
    Max       = Dashboard!$B$6   ($900)
    Parlay    = Dashboard!$B$7   ($25-$50, 2-leg cross-game)

Every other sheet references those cells. Change a number there and it flows
everywhere. Tier dollars are fixed; they do NOT float with bankroll.

Run:  python3 tracker/generate_tracker.py
"""

from datetime import date, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ── LOCKED SIZING — the only sizing numbers in this script ──────────────────
BANKROLL = 3941
BANKROLL_AS_OF = "2026-05-14"
TIER_LEAN = 295
TIER_STANDARD = 510
TIER_MAX = 900
PARLAY_DEFAULT = 50          # range is $25-$50, 2-leg cross-game only
PARLAY_MIN, PARLAY_MAX = 25, 50

DATA_ROWS = 200              # pre-formatted bet rows
FIRST = 2                    # first data row
LAST = FIRST + DATA_ROWS - 1 # 201

SPORTS = ["NBA", "NCAAB", "NFL", "Soccer", "MLB", "NHL", "PGA"]
BET_TYPES = ["Spread", "Over", "Under", "ML", "Parlay"]
TIERS = ["Lean", "Standard", "Max", "Parlay"]
RESULTS = ["W", "L", "P", "Pending"]

# ── Styles ──────────────────────────────────────────────────────────────────
HDR_FILL = PatternFill("solid", fgColor="1F4E2C")
HDR_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(color="1F4E2C", bold=True, size=16)
SUB_FONT = Font(color="666666", italic=True, size=9)
LABEL_FONT = Font(bold=True, size=10)
WARN_FONT = Font(color="9C0006", bold=True, size=10)
MONEY = '"$"#,##0'
PCT = '0.0%'
NUM1 = '0.0'

GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")
GREEN_FONT = Font(color="006100")
RED_FILL = PatternFill("solid", fgColor="FFC7CE")
RED_FONT = Font(color="9C0006")
GREY_FILL = PatternFill("solid", fgColor="D9D9D9")
TIER_FILLS = {
    "Lean": PatternFill("solid", fgColor="D9E1F2"),
    "Standard": PatternFill("solid", fgColor="E2EFDA"),
    "Max": PatternFill("solid", fgColor="FCE4D6"),
    "Parlay": PatternFill("solid", fgColor="E4DFEC"),
}
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=start_col + i, value=h)
        c.fill, c.font = HDR_FILL, HDR_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER


# ── BETS SHEET ────────────────────────────────────────────────────────────--
def build_bets(ws):
    headers = ["Date", "Sport", "Bet Type", "Selection", "Tier", "Stake ($)",
               "Odds (Amer.)", "Bet #", "Close #", "Result", "P&L ($)", "CLV"]
    style_header(ws, 1, headers)
    widths = [12, 9, 10, 28, 11, 10, 12, 9, 9, 9, 11, 9]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = w
    ws.freeze_panes = "A2"

    for r in range(FIRST, LAST + 1):
        # Stake — pulled ONLY from Dashboard tier cells (single source of truth)
        ws.cell(row=r, column=6).value = (
            f'=IF($E{r}="","",IF($E{r}="Lean",Dashboard!$B$4,'
            f'IF($E{r}="Standard",Dashboard!$B$5,'
            f'IF($E{r}="Max",Dashboard!$B$6,'
            f'IF($E{r}="Parlay",Dashboard!$B$7,"")))))'
        )
        # P&L from American odds + result
        ws.cell(row=r, column=11).value = (
            f'=IF($J{r}="","",IF($J{r}="Pending","",IF($J{r}="P",0,'
            f'IF($J{r}="L",-$F{r},IF($G{r}="","",'
            f'IF($G{r}>0,$F{r}*$G{r}/100,$F{r}*100/ABS($G{r})))))))'
        )
        # CLV — direction handled per bet type.
        #   Spread/Under: BetLine - Close   (points)
        #   Over:         Close - BetLine   (points)
        #   ML:           (closeImplied - betImplied) * 100   (probability points)
        ws.cell(row=r, column=12).value = (
            f'=IF(OR($H{r}="",$I{r}=""),"",'
            f'IF($C{r}="Spread",$H{r}-$I{r},'
            f'IF($C{r}="Under",$H{r}-$I{r},'
            f'IF($C{r}="Over",$I{r}-$H{r},'
            f'IF($C{r}="ML",'
            f'(IF($I{r}>0,100/($I{r}+100),ABS($I{r})/(ABS($I{r})+100))'
            f'-IF($H{r}>0,100/($H{r}+100),ABS($H{r})/(ABS($H{r})+100)))*100,'
            f'"")))))'
        )
        ws.cell(row=r, column=6).number_format = MONEY
        ws.cell(row=r, column=11).number_format = MONEY
        ws.cell(row=r, column=12).number_format = NUM1
        ws.cell(row=r, column=1).number_format = "yyyy-mm-dd"
        for col in range(1, 13):
            ws.cell(row=r, column=col).border = BORDER

    # Data validation dropdowns
    dvs = [
        (",".join(SPORTS), "B"),
        (",".join(BET_TYPES), "C"),
        (",".join(TIERS), "E"),
        (",".join(RESULTS), "J"),
    ]
    for opts, col in dvs:
        dv = DataValidation(type="list", formula1=f'"{opts}"', allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"{col}{FIRST}:{col}{LAST}")

    rng = f"{FIRST}:{LAST}"
    # Result coloring
    ws.conditional_formatting.add(f"J{FIRST}:J{LAST}",
        CellIsRule(operator="equal", formula=['"W"'], fill=GREEN_FILL, font=GREEN_FONT))
    ws.conditional_formatting.add(f"J{FIRST}:J{LAST}",
        CellIsRule(operator="equal", formula=['"L"'], fill=RED_FILL, font=RED_FONT))
    ws.conditional_formatting.add(f"J{FIRST}:J{LAST}",
        CellIsRule(operator="equal", formula=['"P"'], fill=GREY_FILL))
    # Tier color-coding
    for tier, fill in TIER_FILLS.items():
        ws.conditional_formatting.add(f"E{FIRST}:E{LAST}",
            CellIsRule(operator="equal", formula=[f'"{tier}"'], fill=fill))
    # P&L + CLV green/red
    for col in ("K", "L"):
        ws.conditional_formatting.add(f"{col}{FIRST}:{col}{LAST}",
            CellIsRule(operator="greaterThan", formula=["0"], font=GREEN_FONT))
        ws.conditional_formatting.add(f"{col}{FIRST}:{col}{LAST}",
            CellIsRule(operator="lessThan", formula=["0"], font=RED_FONT))


# ── DASHBOARD SHEET ──────────────────────────────────────────────────────────
def build_dashboard(ws):
    ws.column_dimensions["A"].width = 22
    for col in "BCDEF":
        ws.column_dimensions[col].width = 13

    ws["A1"] = "BRETT'S BETTING TRACKER"
    ws["A1"].font = TITLE_FONT

    # Bankroll + stale flag
    ws["A2"] = "Bankroll"
    ws["A2"].font = LABEL_FONT
    ws["B2"] = BANKROLL
    ws["B2"].number_format = MONEY
    ws["B2"].font = Font(bold=True, size=12, color="1F4E2C")
    ws["C2"] = f"as of {BANKROLL_AS_OF}"
    ws["C2"].font = SUB_FONT
    ws["D2"] = "STALE — update before sizing"
    ws["D2"].font = WARN_FONT

    # Tier control table (the source of truth the Bets sheet reads)
    ws["A3"] = "TIER CONTROLS"
    ws["A3"].font = LABEL_FONT
    tier_rows = [("Lean", TIER_LEAN), ("Standard", TIER_STANDARD),
                 ("Max", TIER_MAX), ("Parlay", PARLAY_DEFAULT)]
    for i, (name, val) in enumerate(tier_rows):
        r = 4 + i
        ws.cell(row=r, column=1, value=name).font = LABEL_FONT
        c = ws.cell(row=r, column=2, value=val)
        c.number_format, c.font, c.border = MONEY, Font(bold=True), BORDER
        ws.cell(row=r, column=1).border = BORDER
    ws["C7"] = f"range ${PARLAY_MIN}-${PARLAY_MAX}, 2-leg cross-game only"
    ws["C7"].font = SUB_FONT

    b = f"Bets!$K${FIRST}:$K${LAST}"   # P&L range
    j = f"Bets!$J${FIRST}:$J${LAST}"   # Result range
    f = f"Bets!$F${FIRST}:$F${LAST}"   # Stake range
    l = f"Bets!$L${FIRST}:$L${LAST}"   # CLV range
    spt = f"Bets!$B${FIRST}:$B${LAST}" # Sport range
    tr = f"Bets!$E${FIRST}:$E${LAST}"  # Tier range

    # Live performance stats
    ws["A9"] = "PERFORMANCE"
    ws["A9"].font = Font(bold=True, size=12, color="1F4E2C")
    stats = [
        ("Settled bets", f'=COUNTIF({j},"W")+COUNTIF({j},"L")+COUNTIF({j},"P")', "0"),
        ("Pending", f'=COUNTIF({j},"Pending")', "0"),
        ("Record (W-L-P)",
         f'=COUNTIF({j},"W")&"-"&COUNTIF({j},"L")&"-"&COUNTIF({j},"P")', None),
        ("Win %", f'=IFERROR(COUNTIF({j},"W")/(COUNTIF({j},"W")+COUNTIF({j},"L")),"")', PCT),
        ("Total staked",
         f'=SUMIF({j},"W",{f})+SUMIF({j},"L",{f})', MONEY),
        ("Net P&L", f'=SUM({b})', MONEY),
        ("ROI",
         f'=IFERROR(SUM({b})/(SUMIF({j},"W",{f})+SUMIF({j},"L",{f})),"")', PCT),
        ("Avg CLV", f'=IFERROR(AVERAGE({l}),"")', NUM1),
    ]
    for i, (label, formula, fmt) in enumerate(stats):
        r = 10 + i
        ws.cell(row=r, column=1, value=label).font = LABEL_FONT
        c = ws.cell(row=r, column=2, value=formula)
        if fmt:
            c.number_format = fmt
        c.font = Font(bold=True)

    # By-sport breakdown
    ws["A19"] = "BY SPORT"
    ws["A19"].font = Font(bold=True, size=12, color="1F4E2C")
    style_header(ws, 20, ["Sport", "Bets", "Win %", "P&L"])
    for i, sport in enumerate(SPORTS):
        r = 21 + i
        ws.cell(row=r, column=1, value=sport).font = LABEL_FONT
        ws.cell(row=r, column=2,
            value=(f'=COUNTIFS({spt},"{sport}",{j},"W")'
                   f'+COUNTIFS({spt},"{sport}",{j},"L")'
                   f'+COUNTIFS({spt},"{sport}",{j},"P")'))
        ws.cell(row=r, column=3,
            value=(f'=IFERROR(COUNTIFS({spt},"{sport}",{j},"W")/'
                   f'(COUNTIFS({spt},"{sport}",{j},"W")+'
                   f'COUNTIFS({spt},"{sport}",{j},"L")),"")')).number_format = PCT
        ws.cell(row=r, column=4,
            value=f'=SUMIFS({b},{spt},"{sport}")').number_format = MONEY
        for col in range(1, 5):
            ws.cell(row=r, column=col).border = BORDER

    # By-tier breakdown
    ws["A30"] = "BY TIER"
    ws["A30"].font = Font(bold=True, size=12, color="1F4E2C")
    style_header(ws, 31, ["Tier", "Bets", "Win %", "P&L"])
    for i, tier in enumerate(TIERS):
        r = 32 + i
        ws.cell(row=r, column=1, value=tier).font = LABEL_FONT
        ws.cell(row=r, column=2,
            value=(f'=COUNTIFS({tr},"{tier}",{j},"W")'
                   f'+COUNTIFS({tr},"{tier}",{j},"L")'
                   f'+COUNTIFS({tr},"{tier}",{j},"P")'))
        ws.cell(row=r, column=3,
            value=(f'=IFERROR(COUNTIFS({tr},"{tier}",{j},"W")/'
                   f'(COUNTIFS({tr},"{tier}",{j},"W")+'
                   f'COUNTIFS({tr},"{tier}",{j},"L")),"")')).number_format = PCT
        ws.cell(row=r, column=4,
            value=f'=SUMIFS({b},{tr},"{tier}")').number_format = MONEY
        for col in range(1, 5):
            ws.cell(row=r, column=col).border = BORDER
        ws.cell(row=r, column=1).fill = TIER_FILLS[tier]

    # System rules
    ws["A37"] = "SYSTEM RULES"
    ws["A37"].font = Font(bold=True, size=12, color="1F4E2C")
    rules = [
        "Market maker, not a fan — bet the number, not the winner.",
        "Juice ceiling: standard markets -110 to -115. Kill anything > -120.",
        "Softness (softest first): 1H tot > 2H tot > spreads on key # > small dogs (+100..+180) > full-game totals.",
        "CLV is the only weekly metric. +0.3 avg CLV = profitable even in red months.",
        "Cap 5 bets week one. Default LEAN until 10 logged with real CLV.",
        "Log closing line within 1 hr of game start (Pikkit). 5-10 bets/week, weekday cap 3, never chase.",
        f"Sizing: Lean ${TIER_LEAN} / Standard ${TIER_STANDARD} / Max ${TIER_MAX} / Parlay ${PARLAY_MIN}-${PARLAY_MAX} (2-leg cross-game).",
    ]
    for i, rule in enumerate(rules):
        ws.cell(row=38 + i, column=1, value="• " + rule).font = Font(size=10)
        ws.merge_cells(start_row=38 + i, start_column=1, end_row=38 + i, end_column=6)


# ── CHECKLIST SHEET ───────────────────────────────────────────────────────--
def build_checklist(ws):
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 100
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "portrait"
    ws.print_area = "A1:B40"

    def line(r, text, font=None, fill=None, box=False):
        a = ws.cell(row=r, column=1, value="☐" if box else "")
        c = ws.cell(row=r, column=2, value=text)
        if font:
            c.font = font
        if fill:
            c.fill = fill
            a.fill = fill
        if box:
            a.font = Font(size=13)

    line(1, "PRE-BET 4-FILTER GATE", Font(bold=True, size=15, color="1F4E2C"))
    line(2, "All four must pass. Any fail → PASS the bet.", SUB_FONT)

    line(4, "FILTER 1 — Standard market?", LABEL_FONT, GREEN_FILL)
    line(5, "Odds in the -110 to -115 band? KILL anything heavier than -120, "
            "and any parlay/prop/teaser juiced both sides.", box=True)

    line(7, "FILTER 2 — Real edge on the number?", LABEL_FONT, GREEN_FILL)
    line(8, "Projection 1.5+ pts off the number (spread/total) OR 3%+ ML edge.", box=True)

    line(10, "FILTER 3 — Line moved my way?", LABEL_FONT, GREEN_FILL)
    line(11, "Line has moved toward my side (confirmation the market agrees).", box=True)

    line(13, "FILTER 4 — Conviction tier", LABEL_FONT, GREEN_FILL)
    line(14, "Default LEAN ($295). Step up only if 2 of the following are true:", Font(size=10))
    line(15, "Edge 2.5+ pts / 5%+", box=True)
    line(16, "Line moved 1+ pt my way", box=True)
    line(17, "Injury news supports an unpriced angle", box=True)
    line(18, "→ 2 of 3 met = STANDARD ($510). Exceptional + all 3 = MAX ($900).",
         Font(size=10, italic=True))

    line(20, "TIER SIZING (locked)", Font(bold=True, size=12, color="1F4E2C"))
    line(21, f"Lean $295   ·   Standard $510   ·   Max $900   ·   "
             f"Parlay $25-$50 (2-leg cross-game only)", Font(bold=True, size=11))

    line(24, "MARKET-SOFTNESS HIERARCHY (where edge lives — softest first)",
         Font(bold=True, size=12, color="1F4E2C"))
    soft = [
        "1.  1H totals .............. softest, most edge",
        "2.  2H totals",
        "3.  Full-game spreads on key numbers",
        "4.  Small-dog MLs (+100 to +180)",
        "5.  Full-game totals ....... hardest; books model heavily. Only with a pace/efficiency read.",
    ]
    for i, s in enumerate(soft):
        ws.cell(row=25 + i, column=2, value=s).font = Font(size=10)


# ── WEEKLY SHEET ──────────────────────────────────────────────────────────--
def build_weekly(ws):
    headers = ["Week", "Start", "End", "Bets", "P&L ($)", "Avg CLV", "Win %"]
    style_header(ws, 1, headers)
    widths = [7, 12, 12, 8, 12, 10, 9]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = w
    ws.freeze_panes = "A2"

    base = date(2026, 5, 11)  # Monday of the bankroll week
    a = f"Bets!$A${FIRST}:$A${LAST}"
    k = f"Bets!$K${FIRST}:$K${LAST}"
    j = f"Bets!$J${FIRST}:$J${LAST}"
    l = f"Bets!$L${FIRST}:$L${LAST}"

    for w in range(26):
        r = 2 + w
        start = base + timedelta(weeks=w)
        end = start + timedelta(days=6)
        ws.cell(row=r, column=1, value=w + 1).font = LABEL_FONT
        sc = ws.cell(row=r, column=2, value=start)
        ec = ws.cell(row=r, column=3, value=end)
        sc.number_format = ec.number_format = "yyyy-mm-dd"
        ws.cell(row=r, column=4,
            value=f'=COUNTIFS({a},">="&$B{r},{a},"<="&$C{r},{j},"<>Pending",{j},"<>")')
        ws.cell(row=r, column=5,
            value=f'=SUMIFS({k},{a},">="&$B{r},{a},"<="&$C{r})').number_format = MONEY
        ws.cell(row=r, column=6,
            value=f'=IFERROR(AVERAGEIFS({l},{a},">="&$B{r},{a},"<="&$C{r}),"")').number_format = NUM1
        ws.cell(row=r, column=7,
            value=(f'=IFERROR(COUNTIFS({a},">="&$B{r},{a},"<="&$C{r},{j},"W")/'
                   f'(COUNTIFS({a},">="&$B{r},{a},"<="&$C{r},{j},"W")+'
                   f'COUNTIFS({a},">="&$B{r},{a},"<="&$C{r},{j},"L")),"")')).number_format = PCT
        for col in range(1, 8):
            ws.cell(row=r, column=col).border = BORDER

    ws.cell(row=29, column=1, value="TOTAL").font = LABEL_FONT
    ws.cell(row=29, column=5, value=f"=SUM(E2:E27)").number_format = MONEY
    ws.cell(row=29, column=5).font = Font(bold=True)


def main():
    wb = Workbook()
    dash = wb.active
    dash.title = "Dashboard"
    bets = wb.create_sheet("Bets")
    checklist = wb.create_sheet("Checklist")
    weekly = wb.create_sheet("Weekly")

    build_bets(bets)        # build first so Dashboard formulas resolve cleanly
    build_dashboard(dash)
    build_checklist(checklist)
    build_weekly(weekly)

    wb.active = 0
    out = __file__.rsplit("/", 1)[0] + "/bet_tracker.xlsx"
    wb.save(out)
    print(f"Wrote {out}")


if __name__ == "__main__":
    main()
